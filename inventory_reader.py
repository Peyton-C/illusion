from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


class SpreadsheetManager:
    def __init__(self, file_path: str, sheet_name: str = "Inventory", sku_prefix: str = "SKU",):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self.sku_prefix = sku_prefix

        self.sku_header = "SKU"
        self.sku_padding = 6

        self.default_headers = ["SKU", "NAME", "PRIORITY", "ORDER_QUANTITY", "LOW", "LINK_1", "VENDOR_1", "LINK_2", "VENDOR_2", "LINK_3", "VENDOR_3", "LINK_4", "VENDOR_4", "LINK_5", "VENDOR_5",]

        if self.file_path.exists():
            self.workbook = load_workbook(self.file_path)
        else:
            self.workbook = Workbook()

        if self.sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[self.sheet_name]
        else:
            self.sheet = self.workbook.create_sheet(self.sheet_name)
        
        # Validate Headers
        existing_headers = self._get_headers()

        if not existing_headers:
            for column_index, header in enumerate(self.default_headers, start=1):
                self.sheet.cell(row=1, column=column_index).value = header
            return

        for header in self.default_headers:
            if header not in existing_headers:
                next_column = self.sheet.max_column + 1
                self.sheet.cell(row=1, column=next_column).value = header
                existing_headers.append(header)

    def _get_headers(self) -> list[str]:
        headers = []

        for column in range(1, self.sheet.max_column + 1):
            value = self.sheet.cell(row=1, column=column).value

            if value is not None:
                headers.append(str(value))

        return headers
    
    def _get_column_index(self, header_name: str):
        headers = self._get_headers()

        if header_name not in headers:
            raise ValueError(f"Header '{header_name}' does not exist.")

        return headers.index(header_name) + 1
    
    def _row_to_dict(self, row_number: int):
        headers = self._get_headers()
        item = {}

        for column_index, header in enumerate(headers, start=1):
            value = self.sheet.cell(row=row_number, column=column_index,).value
            
            if header == "LOW":
                if value == "TRUE":
                    item[header] = True
                else:
                    item[header] = False

        return item

    def _find_row_by_sku(self, sku: str):
        sku_column = self._get_column_index(self.sku_header)

        for row_number in range(2, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(
                row=row_number,
                column=sku_column,
            ).value

            if str(cell_value) == str(sku):
                return row_number

        return None

    def _generate_sku(self) -> str:
        sku_column = self._get_column_index(self.sku_header)
        highest_number = 0

        for row_number in range(2, self.sheet.max_row + 1):
            sku = self.sheet.cell(row=row_number, column=sku_column).value

            if not sku:
                continue

            sku = str(sku)

            if not sku.startswith(f"{self.sku_prefix}-"):
                continue

            number_part = sku.replace(f"{self.sku_prefix}-", "", 1)

            if number_part.isdigit():
                highest_number = max(highest_number, int(number_part))

        next_number = highest_number + 1
        return f"{self.sku_prefix}-{next_number:0{self.sku_padding}d}"

    def read_all(self) -> list[dict[str, Any]]:
        items = []

        for row_number in range(2, self.sheet.max_row + 1):
            item = self._row_to_dict(row_number)

            if any(value is not None for value in item.values()):
                items.append(item)

        return items
    
    def validate_sku(self, sku: str):
        if self._find_row_by_sku(sku) != None:
            return True
        else:
            return False

    def get_item(self, sku: str):
        row_number = self._find_row_by_sku(sku)

        if row_number is None:
            return None

        return self._row_to_dict(row_number)

    def add_item(self, item_data: dict[str, Any]):
        headers = self._get_headers()
        new_sku = self._generate_sku()

        item_data[self.sku_header] = new_sku

        new_row = []

        for header in headers:
            new_row.append(item_data.get(header))

        self.sheet.append(new_row)

        return new_sku

    def update_item(self, sku: str, updates: dict[str, Any]):
        row_number = self._find_row_by_sku(sku)

        if row_number is None:
            return False

        headers = self._get_headers()

        for header, value in updates.items():
            if header == self.sku_header:
                continue

            if header not in headers:
                raise ValueError(f"Header '{header}' does not exist.")

            column_index = headers.index(header) + 1
            self.sheet.cell(row=row_number, column=column_index).value = value

        return True

    def delete_item(self, sku: str):
        row_number = self._find_row_by_sku(sku)

        if row_number is None:
            return False

        self.sheet.delete_rows(row_number)
        return True
    
    def add_vendor(self, sku: str, vendor_name: str, link: str,):
        item = self.get_item(sku)

        if item is None:
            return False

        for vendor_number in range(1, 6):
            vendor = item.get(f"VENDOR_{vendor_number}")
            existing_link = item.get(f"LINK_{vendor_number}")

            if not vendor and not existing_link:
                self.update_item(sku, {f"VENDOR_{vendor_number}": vendor_name, f"LINK_{vendor_number}": link})
                self.save()

        return False

    def save(self):
        self.workbook.save(self.file_path)